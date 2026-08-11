import openpyxl
import json
import sys

def parse_excel(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Get the first 15 rows to understand the structure
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 15: 
                    break
                rows.append([str(cell) if cell is not None else "" for cell in row])
            result[sheet_name] = rows
            
        with open('output.json', 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
    except Exception as e:
        with open('error.log', 'w', encoding='utf-8') as f:
            f.write(f"Error: {e}")

if __name__ == "__main__":
    parse_excel(sys.argv[1])
